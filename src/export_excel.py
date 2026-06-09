"""Export round-by-round state of KCC-128 encryption (block 1) to Excel."""
from __future__ import annotations
import csv
from pathlib import Path
from knight_chaos_cipher import KnightChaosCipher, _xor_bytes, _rotl, BLOCK_SIZE

KEY = "KCC-24.55.2714-Muhammad-Rizky-Hajar"
PLAINTEXT = "Cyber Security T"  # exactly 16 bytes = 1 block

OUTPUT_DIR = Path(__file__).resolve().parent.parent / "excel"


def fmt_bytes(data: bytes) -> list[str]:
    """Format bytes as uppercase hexadecimal strings for spreadsheet cells."""
    return [f"0x{b:02X}" for b in data]


def main():
    """Export one-block round walkthrough to CSV and XLSX when openpyxl exists."""
    OUTPUT_DIR.mkdir(exist_ok=True)
    cipher = KnightChaosCipher(KEY)
    plain_bytes = PLAINTEXT.encode("utf-8")

    # CBC: first block XORed with IV
    block = _xor_bytes(plain_bytes, cipher.iv)

    rows: list[list[str]] = []
    header = ["Round", "Step", "Penjelasan"] + [f"Byte {i}" for i in range(16)]
    rows.append(header)

    def add_row(rnd, step, explain, state):
        """Append one labeled state row to the walkthrough table."""
        rows.append([str(rnd), step, explain] + fmt_bytes(state))

    add_row(0, "Plaintext (ASCII)", "Teks asli diubah ke byte", plain_bytes)
    add_row(0, "IV", "Initialization Vector dari kunci", cipher.iv)
    add_row(0, "Plaintext XOR IV", "Blok plaintext di-XOR dengan IV (CBC mode)", block)

    # Process all 10 rounds
    state = block
    for rnd_idx, mat in enumerate(cipher.round_material, start=1):
        before = state

        # Step 1: AddRoundKey
        state = _xor_bytes(state, mat.key)
        add_row(rnd_idx, "Round Key", f"Kunci round {rnd_idx}", mat.key)
        add_row(rnd_idx, "AddRoundKey", "state XOR roundKey", state)

        # Step 2: ChaoticSubBytes
        state = bytes(mat.sbox[b] for b in state)
        add_row(rnd_idx, "ChaoticSubBytes", "Ganti tiap byte pakai S-Box dinamis", state)

        # Step 3: KnightPermutation
        perm = mat.permutation
        state_before_perm = state
        new_state = [0] * 16
        for out_i, src_i in enumerate(perm):
            new_state[out_i] = state_before_perm[src_i]
        state = bytes(new_state)
        perm_str = ",".join(str(p) for p in perm)
        add_row(rnd_idx, "KnightPermutation", f"Tukar posisi: [{perm_str}]", state)

        # Step 4: BitShiftMix
        state = bytes(_rotl(b, mat.rotations[i]) for i, b in enumerate(state))
        rot_str = ",".join(str(r) for r in mat.rotations)
        add_row(rnd_idx, "BitShiftMix", f"Rotasi kiri tiap byte: [{rot_str}]", state)

        # Step 5: ByteDiffusion
        state = cipher._diffuse(state, mat.key)
        add_row(rnd_idx, "ByteDiffusion", "XOR berantai: byte[i] ^= carry ^ neighbor ^ key", state)

        # Step 6: FeistelMix
        state = cipher._feistel_mix(state, mat.key)
        add_row(rnd_idx, "FeistelMix", "Mixing berpasangan 3 pass", state)

    add_row("Final", "Ciphertext", "Hasil akhir enkripsi 1 blok", state)

    # Write CSV
    csv_path = OUTPUT_DIR / "kcc128-walkthrough.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerows(rows)

    # Write Excel if openpyxl available
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "KCC-128 Walkthrough"

        for r_idx, row in enumerate(rows, start=1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="333333")
                    cell.alignment = Alignment(horizontal="center")

        # Color code rounds
        colors = {
            "AddRoundKey": "D4E6F1",
            "ChaoticSubBytes": "D5F5E3",
            "KnightPermutation": "FADBD8",
            "BitShiftMix": "FEF9E7",
            "ByteDiffusion": "E8DAEF",
            "FeistelMix": "F9E79F",
            "Round Key": "AED6F1",
        }
        for r_idx, row in enumerate(rows[1:], start=2):
            step = row[1]
            if step in colors:
                fill = PatternFill("solid", fgColor=colors[step])
                for c_idx in range(1, len(row) + 1):
                    ws.cell(row=r_idx, column=c_idx).fill = fill

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

        xlsx_path = OUTPUT_DIR / "kcc128-walkthrough.xlsx"
        wb.save(xlsx_path)
        print(f"Excel saved: {xlsx_path}")
    except ImportError:
        print("openpyxl not installed, skipping .xlsx (CSV still generated)")

    print(f"CSV saved: {csv_path}")
    print(f"\nTotal rows: {len(rows)} (header + {len(rows)-1} data rows)")
    print(f"Covers: Plaintext → IV XOR → 10 rounds × 6 steps → Ciphertext")


if __name__ == "__main__":
    main()
